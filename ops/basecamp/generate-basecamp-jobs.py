#!/usr/bin/env python3
"""Generate Sui Basecamp 2026 t2000 job catalog CSV. Run: python3 ops/basecamp/generate-basecamp-jobs.py"""

import csv
from pathlib import Path
from typing import Optional

HASHTAGS = "#SuiBasecamp #SuiBasecamp2026 #t2000ai @t2000ai @SuiNetwork"
TRUST = "open"  # S.1209 — the ONE trust knob; no tier gates on Basecamp

# (id_suffix, bucket, category, day, zone, title, max_usdc, slots, open_h, sla_h, max_claims, brief, done_when, pack)
JOBS: list[tuple] = []

def add(*row):
    JOBS.append(row)

# --- MICRO ($0.10–0.25) — volume UGC / pulse ---
micro_zones = [
    ("MAINSTAGE", "Summit Mainstage", "Basecamp pulse — 3 openings from Mainstage lobby"),
    ("AI_LAB", "AI Builders Lab", "Basecamp pulse — 3 openings near AI Lab"),
    ("EXCHANGE", "Exchange Stage", "Basecamp pulse — 3 openings at Exchange Stage"),
    ("TRADING", "Trading Arena", "Basecamp pulse — 3 openings at Trading Arena"),
    ("CENTERTREE", "Sui Centerpiece Supertree", "Photo — Sui Centerpiece Supertree"),
    ("CASHGRAB", "Cash Grab Machine", "Photo — Cash Grab Machine in action"),
    ("COFFEE", "Coffee Cart / Anyflo Café", "Photo — coffee moment at Anyflo Café"),
    ("POPCORN", "Popcorn Cart (D1 10:00–12:00)", "Photo — popcorn cart (Day 1 morning only)"),
    ("ICE_CREAM", "Ice Cream Cart (D2 14:00–16:00)", "Photo — ice cream cart (Day 2 afternoon only)"),
    ("STICKER", "Sticker Vault / Sui SWAG", "Photo — your sticker haul at Sticker Vault"),
    ("TOTE", "Tote Studio / DIY Patch", "Photo — customized Basecamp tote"),
    ("AI_SHIRT", "AI T-Shirt Station", "Photo — AI-printed tee"),
    ("AI_PERFUME", "AI Perfume Bar", "Photo — AI fragrance bar"),
    ("SPEAKEASY", "Speakeasy Bar", "Photo — Speakeasy (no spoilers for Hidden Bar location)"),
    ("COMMUNITY", "Community Avenue", "Photo — Community Avenue hangout"),
    ("RING", "ONE Championship Ring", "Photo — ONE ring activation"),
    ("REG", "Registration / Entrance Tunnel", "Photo — entrance tunnel / badge"),
    ("FOOD", "Gathering Table Food Hall", "Photo — food hall moment"),
    ("GORAE", "Gorae EV Charging", "Photo — Gorae EV zone"),
    ("DIY", "DIY Swag Station", "Photo — DIY swag station"),
]

for i, (code, zone, title) in enumerate(micro_zones, 1):
    if "pulse" in title:
        brief = f"At {zone}: call t2000_job_board or check t2000.ai/jobs; quote total/returned; list 3 openings with title+maxUsdc. Include board call timestamp."
        done = "Board quote + 3 openings + Agent ID + timestamp. Same board read cannot split across claims."
        slots = 40
        price = 0.15
    else:
        brief = f"On-site at {zone} (MBS Sands Expo). Take an original photo showing you at the activation; no stock images."
        done = "Photo (link or attach) + one-line caption + Agent ID. Must be visibly Basecamp / MBS context."
        slots = 30
        price = 0.20
    add(f"BC-M-{i:03d}", "micro", "UGC_PHOTO" if "Photo" in title else "PULSE", "D1+D2", zone, title, price, slots, 48, 24, 1, brief, done, "PACK: basecamp-micro")

# Micro X — quick posts
x_micro = [
    ("BC-M-021", "X — quote-tweet t2000 Mainstage moment", "Quote-tweet @t2000ai or @SuiNetwork Mainstage clip with one honest line on agent economy.", 50, 0.15),
    ("BC-M-022", "X — #SuiBasecamp floor check-in", f"Post on X: at MBS, one line what you're building + {HASHTAGS}. Disclosure: paid bounty.", 80, 0.15),
    ("BC-M-023", "X — AI Lab livestream reaction", "Watch AI Builders Lab livestream segment; quote-tweet with one takeaway on agents + MCP.", 40, 0.20),
    ("BC-M-024", "X — photo post with #t2000ai", f"Post your Basecamp photo on X with {HASHTAGS}. Paid disclosure required.", 60, 0.20),
    ("BC-M-025", "X — reply to @t2000ai Basecamp thread", "Reply to the official @t2000ai Basecamp thread with a substantive question or insight (not emoji-only).", 100, 0.10),
]
for jid, title, brief, slots, price in x_micro:
    add(jid, "micro", "X_SOCIAL", "D1+D2", "Anywhere on-site", title, price, slots, 48, 24, 1, brief, "Permalink + screenshot + Agent ID.", "PACK: basecamp-social")

# --- SMALL ($0.30–1.00) ---
small_jobs = [
    ("BC-S-001", "Connect smoke — 2 tools at Basecamp", "MCP", "D1+D2", "t2000 booth / anywhere",
     "Name ≥2 Connect tools used on-site (balance, job_board, agents, job_status). Redacted transcript per tool. AI client + Passport Y/N. Agent ID. Tx digests alone ≠ transcript.", 0.50, 60, 48, 48, 1),
    ("BC-S-002", "Honest friction — one Basecamp blocker", "FRICTION", "D1+D2", "Any zone",
     "One real friction trying hire/claim/deliver/settle OR use MCP at a noisy venue. Evidence: refuse/screenshot/jobId. Expected vs actual ≤3 lines. Agent ID.", 0.75, 40, 48, 48, 1),
    ("BC-S-003", "Mainstage — quote + timestamp", "STAGE_MAIN", "D1", "Summit Mainstage",
     "Attend t2000 Summit Mainstage segment. Deliver: one verbatim quote (≤25 words), session title, approximate timestamp, photo of stage.", 0.50, 50, 24, 24, 1),
    ("BC-S-004", "AI Lab — workshop proof", "STAGE_AI_LAB", "D1+D2", "AI Builders Lab",
     "Attend any AI Builders Lab session mentioning agents/MCP/wallets. Deliver: session name, 2 tool names or demo steps, redacted transcript snippet.", 0.75, 40, 48, 48, 1),
    ("BC-S-005", "Exchange Stage — demo snapshot", "STAGE_EXCHANGE", "D1+D2", "Exchange Stage",
     "Watch an Exchange Stage product demo. Deliver: product name, one thing it does, photo, 3-line summary.", 0.40, 30, 48, 24, 1),
    ("BC-S-006", "Trading Arena — 60s observation", "TRADING", "D1+D2", "Trading Arena",
     "Observe Trading Arena activation 60s+. Deliver: what happened, one metric on screen if visible, photo.", 0.50, 25, 48, 24, 1),
    ("BC-S-007", "Sticker Vault — 5-sticker layout", "UGC", "D1+D2", "Sticker Vault",
     "Collect 5 distinct stickers; flat-lay photo; name each sticker source (project). Agent ID.", 0.35, 35, 48, 24, 1),
    ("BC-S-008", "Tote Studio — patch story", "UGC", "D1+D2", "Tote Studio",
     "Apply one iron-on patch; photo of tote; why you picked that patch (2 lines).", 0.40, 30, 48, 24, 1),
    ("BC-S-009", "Hidden Bar — found it (no location leak)", "SCAVENGER", "D1+D2", "Hidden Bar",
     "Find the Hidden Bar. Deliver: photo INSIDE (no exterior address posted publicly) + one word vibe. Do NOT post location on X — buyer-only delivery.", 1.00, 15, 48, 48, 1),
    ("BC-S-010", "Cash Grab — spectator clip", "UGC", "D1+D2", "Cash Grab Machine",
     "15s video or 3 photos of Cash Grab attempt (faces optional). No wallet addresses on screen.", 0.50, 20, 48, 24, 1),
    ("BC-S-011", "Register Agent ID at Basecamp", "ONBOARD", "D1+D2", "Registration / booth",
     "Register new Agent ID via Connect at venue. Deliver: #id, profile URL, redacted t2000_agents transcript. Must be fresh (created during event window).", 0.50, 80, 48, 72, 1),
    ("BC-S-012", "X thread — 3 posts Basecamp Day 1", "X_SOCIAL", "D1", "Anywhere",
     f"Thread ≥3 posts on X covering Day 1; mention t2000 once; {HASHTAGS}; paid disclosure each post.", 0.80, 25, 24, 24, 1),
    ("BC-S-013", "X thread — 3 posts Basecamp Day 2", "X_SOCIAL", "D2", "Anywhere",
     f"Thread ≥3 posts Day 2; include AI Lab or Mainstage beat; {HASHTAGS}.", 0.80, 25, 24, 24, 1),
    ("BC-S-014", "Booth scan — t2000 QR", "BOOTH_T2000", "D1+D2", "t2000 booth",
     "Visit t2000 booth; scan QR; complete one Connect action. Deliver: which flow + redacted result.", 0.45, 50, 48, 24, 1),
    ("BC-S-015", "Cross-booth — one ecosystem hire", "CROSS", "D1+D2", "Expo floor",
     "Hire or claim ONE non-t2000 Open or Service from a booth partner (≥$0.10). Deliver: jobId + title + one-line outcome.", 0.60, 30, 48, 48, 1),
]
for row in small_jobs:
    jid, title, cat, day, zone, brief, price, slots, oh, sla, mc = row
    done = "Per brief; UNIQUE PROOF per job; Agent ID required."
    add(jid, "small", cat, day, zone, title, price, slots, oh, sla, mc, brief, done, "PACK: basecamp-small")

# --- MED ($1.00–5.00) ---
med_jobs = [
    ("BC-MED-001", "Job loop — hire a stranger at Basecamp", "JOB_LOOP", "D1+D2", "Expo floor",
     "YOU fund proof job ≥$0.25 USDC with substantive task (not one-liner). Different seller delivers; YOU settle; deliver proof jobId + both Agent IDs + path. No desk bounty as proof. No repeat proof seller.", 2.50, 40, 48, 72, 1),
    ("BC-MED-002", "Refer — onboard agent at Basecamp NO Path A", "REFERRAL", "D1+D2", "Expo floor",
     "Refer NEW agent (#id); their FIRST released seller job is proof (releasedCount=1); hunter≠buyer on proof; both IDs in delivery. NOT hunter-funded micro hire.", 3.00, 25, 48, 72, 1),
    ("BC-MED-003", "Mainstage — live notes (10 bullets)", "STAGE_MAIN", "D1", "Summit Mainstage",
     "Attend full t2000 Mainstage talk. Deliver 10 bullet notes + 1 photo + 1 actionable takeaway for builders.", 2.00, 20, 24, 48, 1),
    ("BC-MED-004", "AI Lab — build with MCP (15 min)", "STAGE_AI_LAB", "D1+D2", "AI Builders Lab",
     "15+ min hands-on with MCP at Lab. Deliver: goal, tools used, redacted transcripts, screenshot, what shipped.", 3.00, 20, 48, 72, 1),
    ("BC-MED-005", "Video — 30s t2000 hype reel", "CONTENT", "D1+D2", "Any on-site",
     "30–60s vertical video: you using t2000 or explaining agent marketplace. MBS b-roll OK. Link + thumbnail.", 2.50, 15, 48, 72, 1),
    ("BC-MED-006", "Scavenger — 5 zones in 90 min", "SCAVENGER", "D1+D2", "Multi-zone",
     "Photo proof at 5 distinct zones (list in delivery): Mainstage, AI Lab, Centerpiece, SWAG, one sponsor booth. Timestamp each.", 1.50, 30, 48, 48, 1),
    ("BC-MED-007", "Pair loop — two agents one deliverable", "JOB_LOOP", "D1+D2", "Expo floor",
     "Pair with another hunter; one posts Open ≥$0.50, other delivers, poster settles. Both claim this bounty with SAME proof jobId + both hunter IDs.", 3.50, 15, 48, 72, 1),
    ("BC-MED-008", "Friction deep-dive — venue + product", "FRICTION", "D1+D2", "Any",
     "Reproduce one bug OR document one UX gap with steps, evidence, severity, suggested fix. Mainstage/Lab/booth context preferred.", 2.00, 25, 48, 72, 1),
    ("BC-MED-009", "Livestream clip — AI Lab timestamp", "LIVESTREAM", "D1+D2", "AI Builders Lab",
     "Clip from official livestream with timestamp + URL; explain what demo showed in 5 lines.", 1.25, 20, 48, 48, 1),
    ("BC-MED-010", "Diamond booth crawl — 3 demos", "CROSS", "D1+D2", "Diamond row",
     "Visit 3 Diamond booths; one photo each; one sentence per demo; tag projects.", 1.75, 20, 48, 48, 1),
    ("BC-MED-011", "Night mode — Speakeasy / Afterparty", "AFTER", "D1 eve", "Speakeasy",
     "Evening activation: photo + 3-line story of one conversation about agents/work. No private info.", 2.00, 15, 36, 48, 1),
    ("BC-MED-012", "X — quote card graphic", "X_SOCIAL", "D1+D2", "Any",
     f"Design simple quote card from Mainstage/Lab line; post on X with {HASHTAGS}.", 1.50, 20, 48, 48, 1),
]
for row in med_jobs:
    jid, title, cat, day, zone, brief, price, slots, oh, sla, mc = row
    add(jid, "med", cat, day, zone, title, price, slots, oh, sla, mc, brief, "Per brief; UNIQUE PROOF; anti-self-deal.", "PACK: basecamp-med")

# --- LARGE ($5–25) ---
large_jobs = [
    ("BC-L-001", "Ambassador — full Day 1 recap article", "CONTENT", "D1", "Multi",
     "500–800 word Day 1 recap: Mainstage + 2 activations + t2000 angle. Publish Medium/blog/X longform. Link.", 15.00, 5, 36, 72, 1),
    ("BC-L-002", "Ambassador — full Day 2 recap article", "CONTENT", "D2", "Multi",
     "500–800 word Day 2 recap: AI Lab + floor + t2000. Link.", 15.00, 5, 36, 72, 1),
    ("BC-L-003", "Video — 3min documentary style", "CONTENT", "D1+D2", "Multi",
     "3–5 min edit: your Basecamp arc + t2000 mention ≥30s. YouTube/X video link.", 20.00, 4, 48, 96, 1),
    ("BC-L-004", "Referral chain — 3 agents onboarded", "REFERRAL", "D1+D2", "Expo",
     "Three distinct referred agents each with first released seller job; ledger-unique; NO Path A. Table of IDs + jobIds.", 18.00, 3, 48, 96, 1),
    ("BC-L-005", "Job-loop marathon — 3 settled proofs", "JOB_LOOP", "D1+D2", "Expo",
     "Three separate proof jobs you funded+settled (≥$0.25 each, distinct sellers). Deliver 3 proof jobIds + timeline.", 12.00, 5, 48, 96, 1),
    ("BC-L-006", "Mainstage — official thread takeover", "STAGE_MAIN", "D1", "Mainstage",
     "Live-tweet Mainstage talk: ≥8 tweets, timestamps, 2 photos, quote tweets. Thread link.", 10.00, 6, 24, 48, 1),
    ("BC-L-007", "AI Lab — ship micro-tool with MCP", "STAGE_AI_LAB", "D1+D2", "AI Lab",
     "Ship tiny open-source or gist tool using t2000 MCP; repo link + 2min demo video + README.", 25.00, 3, 48, 96, 1),
    ("BC-L-008", "Booth duty — 2h t2000 evangelist", "BOOTH_T2000", "D1+D2", "t2000 booth",
     "Staff booth 2h (pre-coordinated slot); deliver sign-in photo hourly + count of Connect setups helped (honor system + 3 testimonials).", 12.00, 8, 48, 48, 1),
    ("BC-L-009", "Podcast — 15min interview clip", "CONTENT", "D1+D2", "Quiet zone",
     "Record 15min interview with another attendee about agent economies; publish clip; link.", 15.00, 4, 48, 96, 1),
    ("BC-L-010", "Photo essay — 12 images", "UGC", "D1+D2", "Full floor",
     "12-photo essay with captions covering floorplan highlights; publish X thread or blog.", 8.00, 6, 48, 72, 1),
]
for row in large_jobs:
    jid, title, cat, day, zone, brief, price, slots, oh, sla, mc = row
    add(jid, "large", cat, day, zone, title, price, slots, oh, sla, mc, brief, "Per brief; highest quality bar.", "PACK: basecamp-large")

# Expand micro with zone-specific variants (duplicate templates, more slots) — hit volume
extras = [
    ("BC-M-026", "micro", "PULSE", "D1", "Summit Mainstage", "D1 AM pulse — board before noon", 0.15, 30, 12, 12, 1,
     "Before 12:00 SGT Day 1: board pulse + 3 openings.", "Timestamp before noon SGT.", "PACK: basecamp-micro"),
    ("BC-M-027", "micro", "PULSE", "D2", "AI Builders Lab", "D2 PM pulse — board after 14:00", 0.15, 30, 12, 12, 1,
     "After 14:00 SGT Day 2: board pulse + 3 openings.", "Timestamp after 14:00 SGT.", "PACK: basecamp-micro"),
    ("BC-M-028", "micro", "HYPE", "D1", "Mainstage queue", "Hype — Mainstage queue photo", 0.20, 40, 24, 12, 1,
     "Photo in Mainstage queue or entrance line; one line why you're here.", "Photo + caption.", "PACK: basecamp-micro"),
    ("BC-M-029", "micro", "HYPE", "D2", "AI Lab", "Hype — AI Lab full room", 0.20, 40, 24, 12, 1,
     "Photo showing AI Lab attendance; no fake crowd.", "Photo + session name if visible.", "PACK: basecamp-micro"),
    ("BC-M-030", "micro", "X_SOCIAL", "D1+D2", "Any", "X — bookmark + repost t2000 clip", 0.10, 100, 48, 24, 1,
     "Repost official t2000 Basecamp clip with one sentence takeaway.", "Permalink.", "PACK: basecamp-social"),
]
for e in extras:
    add(*e)

# --- Booth grid — ~35 activations × micro photo (scale to thousands) ---
for n in range(1, 36):
    add(
        f"BC-M-B{n:03d}", "micro", "BOOTH_CRAWL", "D1+D2", f"Expo booth #{n}",
        f"Basecamp booth visit #{n} — photo + one-line pitch",
        0.15, 25, 48, 24, 1,
        f"Visit sponsor/ecosystem booth #{n} on floorplan row; photo with booth branding; one sentence what they do. Agent ID.",
        "Photo + booth name + Agent ID. UNIQUE per booth number.",
        "PACK: basecamp-booth",
    )

# --- Hourly hype windows (event 11:00–17:00 × 2 days = 12h/day) ---
for day, dlabel in [(1, "D1"), (2, "D2")]:
    for hour in range(11, 18):
        add(
            f"BC-M-H{day}{hour}", "micro", "HYPE", dlabel, "Floor-wide",
            f"{dlabel} {hour}:00 SGT — floor temperature check",
            0.12, 15, 6, 6, 1,
            f"During {dlabel} ~{hour}:00–{hour+1}:00 SGT: one photo of crowd energy + estimated headcount band (sparse/medium/packed).",
            f"Photo timestamp within window; Agent ID.",
            "PACK: basecamp-hype",
        )

# --- X engagement ladder (volume) ---
for n in range(1, 21):
    add(
        f"BC-M-X{n:02d}", "micro", "X_SOCIAL", "D1+D2", "Any",
        f"X engagement #{n} — reply with insight (not spam)",
        0.10, 50, 48, 24, 1,
        f"Find a Basecamp-related post (t2000/Sui/builders); reply with ≥2 sentences of substance. Disclosure: paid bounty. Variation #{n}.",
        "Permalink to your reply + Agent ID.",
        "PACK: basecamp-social",
    )

# --- Med duplicates with zone variants ---
zones_med = ["Trading Arena", "Coffee Cart", "Popcorn Cart", "Ice Cream Cart", "Cash Grab", "Tote Studio"]
for i, z in enumerate(zones_med, 1):
    add(
        f"BC-MED-Z{i}", "med", "JOB_LOOP", "D1+D2", z,
        f"Job loop — hire at {z} (≥$0.25 proof)",
        2.50, 12, 48, 72, 1,
        f"Fund proof job ≥$0.25 near {z}; different seller; you settle; full loop proof. No penny proofs.",
        "proof jobId + both Agent IDs + released status.",
        "PACK: basecamp-med",
    )

# --- Small — MCP at each stage ---
for stage, code in [("Mainstage", "MAIN"), ("AI Lab", "LAB"), ("Exchange", "EXCH")]:
    add(
        f"BC-S-MCP-{code}", "small", "MCP_CONNECT", "D1+D2", stage,
        f"MCP demo — 3 tools from {stage}",
        0.65, 35, 48, 48, 1,
        f"At/near {stage}: call 3 Connect tools; redacted transcripts; timestamp; Agent ID.",
        "3 tool names + transcripts.",
        "PACK: basecamp-small",
    )

OUT_DIR = Path(__file__).parent
OUT_CSV = OUT_DIR / "SUI-BASECAMP-2026-JOBS.csv"
OUT_XLSX = OUT_DIR / "SUI-BASECAMP-2026-JOBS.xlsx"

# Full export (All jobs tab + CSV)
FIELDS = [
    "job_id", "bucket", "category", "day", "zone", "title", "max_usdc", "slots",
    "est_escrow_usdc", "open_hours", "sla_hours", "max_claims_per_agent", "trust_requirement",
    "hashtags", "pack_tag", "brief", "done_when",
]

# Per-bucket tabs — omit redundant `bucket` column for readability
SHEET_FIELDS = [
    "job_id", "category", "day", "zone", "title", "max_usdc", "slots",
    "est_escrow_usdc", "open_hours", "sla_hours", "max_claims_per_agent",
    "trust_requirement", "pack_tag", "brief", "done_when",
]

SHEET_HEADERS = {
    "job_id": "Job ID",
    "category": "Category",
    "day": "Day",
    "zone": "Zone / Location",
    "title": "Open job title",
    "max_usdc": "Max USDC",
    "slots": "Batch slots",
    "est_escrow_usdc": "Est. escrow (max)",
    "open_hours": "Open hours",
    "sla_hours": "SLA hours",
    "max_claims_per_agent": "Max claims / agent",
    "trust_requirement": "Trust requirement (open=Anyone)",
    "pack_tag": "PACK tag",
    "brief": "Brief (post verbatim)",
    "done_when": "Done when",
    "bucket": "Bucket",
    "hashtags": "Hashtags",
}

BUDGET_ROWS = [
    ("lean_1k", 1000, 0.35, "micro + X social", "~1,500", 950, "Photos + X only; skip large"),
    ("standard_5k", 5000, 2.2, "all buckets", "~9,500", 5000, "Full catalog; multiply slots ×2.2"),
    ("max_10k", 10000, 4.5, "all buckets", "~19,500", 10000, "Full catalog; multiply slots ×4.5"),
    ("mainstage_focus", 3000, 1.0, "stage + booth + MCP", "~800", 2800, "Quality over booth-crawl volume"),
    ("ugc_wall", 2000, 3.0, "UGC + hype + X", "~6,000", 2000, "Social / photo wall"),
]

rows_out = []
total_escrow = 0
total_slots = 0
for j in JOBS:
    (jid, bucket, cat, day, zone, title, price, slots, oh, sla, mc, brief, done, pack) = j
    esc = round(price * slots, 2)
    total_escrow += esc
    total_slots += slots
    rows_out.append({
        "job_id": jid,
        "bucket": bucket,
        "category": cat,
        "day": day,
        "zone": zone,
        "title": title,
        "max_usdc": price,
        "slots": slots,
        "est_escrow_usdc": esc,
        "open_hours": oh,
        "sla_hours": sla,
        "max_claims_per_agent": mc,
        "trust_requirement": TRUST,
        "hashtags": HASHTAGS,
        "pack_tag": pack,
        "brief": brief,
        "done_when": done,
    })

with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=FIELDS)
    w.writeheader()
    w.writerows(rows_out)

# --- Excel workbook (tabs per bucket) ---
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Bucket palette — header / zebra / tab (t2000-friendly accents)
BUCKET_THEME = {
    "micro": {
        "header": "0E7490",
        "header_font": "FFFFFF",
        "stripe_a": "ECFEFF",
        "stripe_b": "FFFFFF",
        "subtotal": "A5F3FC",
        "tab": "06B6D4",
    },
    "small": {
        "header": "047857",
        "header_font": "FFFFFF",
        "stripe_a": "ECFDF5",
        "stripe_b": "FFFFFF",
        "subtotal": "A7F3D0",
        "tab": "10B981",
    },
    "med": {
        "header": "B45309",
        "header_font": "FFFFFF",
        "stripe_a": "FFFBEB",
        "stripe_b": "FFFFFF",
        "subtotal": "FDE68A",
        "tab": "F59E0B",
    },
    "large": {
        "header": "6D28D9",
        "header_font": "FFFFFF",
        "stripe_a": "F5F3FF",
        "stripe_b": "FFFFFF",
        "subtotal": "DDD6FE",
        "tab": "8B5CF6",
    },
}

# Category chips (category column tint on data rows)
CATEGORY_FILL = {
    "UGC_PHOTO": "E0E7FF",
    "PULSE": "DBEAFE",
    "X_SOCIAL": "FCE7F3",
    "FRICTION": "FEE2E2",
    "JOB_LOOP": "D1FAE5",
    "REFERRAL": "FEF3C7",
    "STAGE_MAIN": "EDE9FE",
    "STAGE_AI_LAB": "E9D5FF",
    "STAGE_EXCHANGE": "F3E8FF",
    "MCP_CONNECT": "CCFBF1",
    "BOOTH_T2000": "CFFAFE",
    "SCAVENGER": "FFEDD5",
    "CONTENT": "F1F5F9",
    "HYPE": "FFF7ED",
    "BOOTH_CRAWL": "F8FAFC",
    "TRADING": "E2E8F0",
    "ONBOARD": "DCFCE7",
    "CROSS": "F1F5F9",
    "LIVESTREAM": "E0F2FE",
    "AFTER": "1E293B",
    "UGC": "E0E7FF",
}

OVERVIEW_HEADER = PatternFill("solid", fgColor="18181B")
NEUTRAL_HEADER = PatternFill("solid", fgColor="27272A")
TITLE_FONT = Font(bold=True, size=16, color="18181B")
SECTION_FONT = Font(bold=True, size=12, color="18181B")
LABEL_FONT = Font(bold=True, color="52525B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=10, color="18181B")
SUBTOTAL_FONT = Font(bold=True, size=10, color="18181B")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="E4E4E7")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_USD = "$#,##0.00"
FMT_INT = "#,##0"

COL_WIDTHS = {
    "job_id": 14, "category": 18, "day": 10, "zone": 30, "title": 44,
    "max_usdc": 11, "slots": 11, "est_escrow_usdc": 15,
    "open_hours": 11, "sla_hours": 10, "max_claims_per_agent": 14,
    "trust_requirement": 16, "pack_tag": 20, "brief": 58, "done_when": 42,
    "bucket": 10, "hashtags": 38,
}

LONG_TEXT_KEYS = frozenset({"brief", "done_when", "title", "zone", "hashtags"})
MONEY_KEYS = frozenset({"max_usdc", "est_escrow_usdc"})
INT_KEYS = frozenset({"slots", "open_hours", "sla_hours", "max_claims_per_agent"})


INPUT_FILL = PatternFill("solid", fgColor="FEF9C3")
INPUT_FONT = Font(bold=True, size=11, color="854D0E")


def col_letter(fields: list[str], key: str) -> str:
    return get_column_letter(fields.index(key) + 1)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def style_header_row(ws, fields: list[str], theme: Optional[dict], start_row: int = 1):
    hdr_fill = fill(theme["header"]) if theme else OVERVIEW_HEADER
    hdr_font = Font(bold=True, color=(theme.get("header_font", "FFFFFF") if theme else "FFFFFF"), size=10)
    for col_idx, key in enumerate(fields, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=SHEET_HEADERS.get(key, key))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = CENTER
        cell.border = GRID
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(key, 14)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(fields))}{start_row}"
    ws.row_dimensions[start_row].height = 32


def write_job_rows(
    ws,
    fields: list[str],
    data: list[dict],
    theme: Optional[dict],
    start_row: int = 2,
    bucket_key: Optional[str] = "bucket",
):
    cat_col = fields.index("category") + 1 if "category" in fields else None
    slots_col = col_letter(fields, "slots")
    usdc_col = col_letter(fields, "max_usdc")
    esc_col = col_letter(fields, "est_escrow_usdc")
    last_data_row = start_row + len(data) - 1 if data else start_row

    for r_off, row in enumerate(data):
        r_idx = start_row + r_off
        if theme:
            stripe = theme["stripe_a"] if r_off % 2 == 0 else theme["stripe_b"]
        else:
            stripe = "FAFAFA" if r_off % 2 == 0 else "FFFFFF"
        row_fill = fill(stripe)
        ws.row_dimensions[r_idx].height = 72 if len(str(row.get("brief", ""))) > 120 else 48
        for c_idx, key in enumerate(fields, 1):
            if key == "est_escrow_usdc":
                cell = ws.cell(row=r_idx, column=c_idx, value=f"={usdc_col}{r_idx}*{slots_col}{r_idx}")
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
            cell.font = BODY_FONT
            cell.border = GRID
            cell.alignment = WRAP if key in LONG_TEXT_KEYS else TOP
            if key == "max_usdc":
                cell.number_format = FMT_USD
            elif key == "est_escrow_usdc":
                cell.number_format = FMT_USD
            elif key in INT_KEYS:
                cell.number_format = FMT_INT
            if cat_col and c_idx == cat_col:
                cat = str(row.get("category", ""))
                cell.fill = fill(CATEGORY_FILL.get(cat, stripe))
            elif bucket_key and key == "bucket":
                b = str(row.get("bucket", ""))
                t = BUCKET_THEME.get(b)
                cell.fill = fill(t["stripe_a"]) if t else row_fill
                cell.font = Font(bold=True, size=10, color=t["header"] if t else "18181B")
            else:
                cell.fill = row_fill

    return last_data_row


def write_subtotal(ws, fields: list[str], data: list[dict], row: int, theme: dict, data_start: int = 2):
    sub_fill = fill(theme["subtotal"])
    slots_i = fields.index("slots") + 1
    esc_i = fields.index("est_escrow_usdc") + 1
    slots_l = get_column_letter(slots_i)
    esc_l = get_column_letter(esc_i)
    last = row - 1
    for c in range(1, len(fields) + 1):
        ws.cell(row=row, column=c).fill = sub_fill
        ws.cell(row=row, column=c).border = GRID
    ws.cell(row=row, column=1, value="SUBTOTAL").font = SUBTOTAL_FONT
    ws.cell(row=row, column=slots_i, value=f"=SUM({slots_l}{data_start}:{slots_l}{last})").font = SUBTOTAL_FONT
    ws.cell(row=row, column=slots_i).number_format = FMT_INT
    ws.cell(row=row, column=esc_i, value=f"=SUM({esc_l}{data_start}:{esc_l}{last})").font = SUBTOTAL_FONT
    ws.cell(row=row, column=esc_i).number_format = FMT_USD
    ws.row_dimensions[row].height = 24
    return row


wb = Workbook()
wb.remove(wb.active)

by_bucket: dict[str, dict] = {}
for row in rows_out:
    b = row["bucket"]
    by_bucket.setdefault(b, {"templates": 0, "slots": 0, "escrow": 0.0, "prices": []})
    by_bucket[b]["templates"] += 1
    by_bucket[b]["slots"] += row["slots"]
    by_bucket[b]["escrow"] += row["est_escrow_usdc"]
    by_bucket[b]["prices"].append(row["max_usdc"])

bucket_order = ["micro", "small", "med", "large"]

# --- Bucket tabs ---
bucket_subrows: dict[str, int] = {}
for bucket in bucket_order:
    theme = BUCKET_THEME[bucket]
    ws = wb.create_sheet(bucket)
    ws.sheet_properties.tabColor = theme["tab"]
    bucket_rows = [r for r in rows_out if r["bucket"] == bucket]
    style_header_row(ws, SHEET_FIELDS, theme)
    write_job_rows(ws, SHEET_FIELDS, bucket_rows, theme, bucket_key=None)
    sub_row = len(bucket_rows) + 2
    write_subtotal(ws, SHEET_FIELDS, bucket_rows, sub_row, theme)
    bucket_subrows[bucket] = sub_row

# --- Planner (dynamic budget) ---
pl = wb.create_sheet("Planner")
pl.sheet_properties.tabColor = "EAB308"
pl["A1"] = "Budget planner"
pl["A1"].font = TITLE_FONT
pl["A2"] = "Target budget (B4) is your goal. Scaled total hits it when Slot × follows E4 (default). Gap = target − scaled."
pl["A2"].font = Font(size=10, color="71717A")
pl.merge_cells("A1:G1")
pl.merge_cells("A2:G2")

pl["A4"] = "Target budget"
pl["A4"].font = LABEL_FONT
tgt = pl["B4"]
tgt.value = 5000
tgt.fill = INPUT_FILL
tgt.font = INPUT_FONT
tgt.number_format = FMT_USD
tgt.border = GRID

pl["D4"] = "Suggested global Slot ×"
pl["D4"].font = LABEL_FONT
# E4 formula set after tot_r is known — placeholder, patched below

hdr_row = 6
pl_headers = ["Bucket", "Templates", "Base slots", "Base escrow", "Slot ×", "Scaled slots", "Scaled escrow"]
for c, h in enumerate(pl_headers, 1):
    cell = pl.cell(row=hdr_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = NEUTRAL_HEADER
    cell.alignment = CENTER
    cell.border = GRID
pl.freeze_panes = "A7"

slots_l = col_letter(SHEET_FIELDS, "slots")
esc_l = col_letter(SHEET_FIELDS, "est_escrow_usdc")
planner_start = hdr_row + 1
for i, bucket in enumerate(bucket_order):
    r = planner_start + i
    if bucket not in bucket_subrows:
        continue
    sub = bucket_subrows[bucket]
    theme = BUCKET_THEME[bucket]
    n_tpl = by_bucket[bucket]["templates"]
    for c in range(1, 8):
        pl.cell(row=r, column=c).fill = fill(theme["stripe_a"])
        pl.cell(row=r, column=c).border = GRID
    pl.cell(row=r, column=1, value=bucket.upper()).font = Font(bold=True, color=theme["header"])
    pl.cell(row=r, column=2, value=n_tpl).font = BODY_FONT
    pl.cell(row=r, column=3, value=f"={bucket}!{slots_l}{sub}")
    pl.cell(row=r, column=3).number_format = FMT_INT
    pl.cell(row=r, column=4, value=f"={bucket}!{esc_l}{sub}")
    pl.cell(row=r, column=4).number_format = FMT_USD
    mult = pl.cell(row=r, column=5, value="=$E$4")
    mult.fill = fill("FEF9C3")
    mult.font = Font(size=10, color="854D0E")
    mult.number_format = "0.00×"
    mult.border = GRID
    pl.cell(row=r, column=6, value=f"=C{r}*E{r}")
    pl.cell(row=r, column=6).number_format = FMT_INT
    pl.cell(row=r, column=7, value=f"=D{r}*E{r}")
    pl.cell(row=r, column=7).number_format = FMT_USD

tot_r = planner_start + len(bucket_order)
for c in range(1, 8):
    pl.cell(row=tot_r, column=c).fill = fill("F4F4F5")
    pl.cell(row=tot_r, column=c).border = GRID
pl.cell(row=tot_r, column=1, value="TOTAL (scaled)").font = SUBTOTAL_FONT
pl.cell(row=tot_r, column=6, value=f"=SUM(F{planner_start}:F{tot_r-1})").font = SUBTOTAL_FONT
pl.cell(row=tot_r, column=6).number_format = FMT_INT
pl.cell(row=tot_r, column=7, value=f"=SUM(G{planner_start}:G{tot_r-1})").font = SUBTOTAL_FONT
pl.cell(row=tot_r, column=7).number_format = FMT_USD

gap_r = tot_r + 1
pl.cell(row=gap_r, column=1, value="Gap vs target").font = LABEL_FONT
pl.cell(row=gap_r, column=7, value=f"=B4-G{tot_r}")
pl.cell(row=gap_r, column=7).number_format = FMT_USD
pl.cell(row=gap_r, column=7).font = Font(bold=True, size=11, color="B45309")

base_escrow_sum = f"SUM(D{planner_start}:D{tot_r - 1})"
pl["E4"] = f"=IFERROR(B4/{base_escrow_sum},\"—\")"
pl["E4"].number_format = "0.00×"
pl["E4"].font = Font(bold=True, size=11, color="047857")
pl["F4"] = "← global Slot ×; bucket column E auto-follows (type a number to override one row)"
pl["F4"].font = Font(italic=True, size=9, color="71717A")
pl.merge_cells("F4:G4")

how = gap_r + 2
pl.cell(row=how, column=1, value="How to scale").font = SECTION_FONT
how_lines = [
    "1. Base catalog ≈ $2.3k / 4.3k slots at Slot × = 1 (before any scaling).",
    "2. Set Target budget (B4) → Suggested global × (E4) = target ÷ base escrow.",
    "3. Bucket Slot × (column E) default = E4 — scaled total should match target (gap ≈ $0).",
    "4. Override one bucket's E cell with a number (e.g. 3) to weight micro vs large.",
    "5. Edit Max USDC / Batch slots on bucket tabs → base + scaled recalc live.",
    "6. Scaled escrow = max posting cost if every slot fills (before 5% settle fee).",
]
for j, line in enumerate(how_lines, how + 1):
    pl.cell(row=j, column=1, value=line).font = Font(size=10, color="52525B")
    pl.merge_cells(start_row=j, start_column=1, end_row=j, end_column=7)

for i, w in enumerate([16, 11, 12, 14, 10, 13, 14], 1):
    pl.column_dimensions[get_column_letter(i)].width = w

# --- Budget (static reference) ---
bud = wb.create_sheet("Budget")
bud.sheet_properties.tabColor = "71717A"
bud["A1"] = "Budget scenarios"
bud["A1"].font = TITLE_FONT
bud["A2"] = "Static reference only — use Planner tab for live scaling."
bud["A2"].font = Font(size=10, color="71717A")
bud.merge_cells("A1:G1")
bud.merge_cells("A2:G2")
bud_headers = ["Scenario", "Target $", "Slot ×", "Categories", "Est. jobs", "Est. escrow $", "Notes"]
for c, h in enumerate(bud_headers, 1):
    cell = bud.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = NEUTRAL_HEADER
    cell.alignment = CENTER
    cell.border = GRID
bud.freeze_panes = "A5"
budget_fills = ["F0FDF4", "FFFFFF", "FFFBEB", "F5F3FF", "F8FAFC"]
for r, (row, bg) in enumerate(zip(BUDGET_ROWS, budget_fills), 5):
    for c, val in enumerate(row, 1):
        cell = bud.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        cell.fill = fill(bg)
        cell.border = GRID
        cell.alignment = WRAP if c == 7 else TOP
        if c == 2 or c == 6:
            cell.number_format = FMT_USD
        if c == 3:
            cell.number_format = "0.00×"
for i, w in enumerate([20, 11, 10, 24, 12, 14, 50], 1):
    bud.column_dimensions[get_column_letter(i)].width = w
bud.row_dimensions[4].height = 28

# --- All jobs ---
all_ws = wb.create_sheet("All jobs")
all_ws.sheet_properties.tabColor = "A1A1AA"
style_header_row(all_ws, FIELDS, None)
write_job_rows(all_ws, FIELDS, rows_out, None, bucket_key="bucket")
write_subtotal(all_ws, FIELDS, rows_out, len(rows_out) + 2, {
    "subtotal": "E4E4E7",
    "stripe_a": "FAFAFA",
})

# --- Overview (landing — live links to Planner, no duplicate bucket math) ---
ov = wb.create_sheet("Overview", 0)
ov.sheet_properties.tabColor = "18181B"
ov["A1"] = "Sui Basecamp 2026"
ov["A1"].font = Font(bold=True, size=20, color="18181B")
ov["A2"] = "t2000 Open job catalog — start on Planner for budget; bucket tabs for job copy"
ov["A2"].font = Font(size=11, color="71717A")
ov.merge_cells("A1:F1")
ov.merge_cells("A2:F2")

info = [
    ("Event", "Wed 7 Oct – Thu 8 Oct 2026 · 11:00–17:00 SGT"),
    ("Venue", "Marina Bay Sands Expo & Convention Centre, Singapore"),
    ("t2000", "Summit Mainstage + AI Builders Lab (livestreamed)"),
    ("Hashtags", HASHTAGS),
    ("Posting", "trustRequirement open · no tier gates · job-loop proof ≥ $0.25"),
    ("Regenerate", "python3 ops/basecamp/generate-basecamp-jobs.py"),
]
for i, (k, v) in enumerate(info, 4):
    ov.cell(row=i, column=1, value=k).font = LABEL_FONT
    c = ov.cell(row=i, column=2, value=v)
    c.font = BODY_FONT
    c.alignment = WRAP
    ov.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

snap = 11
ov.cell(row=snap, column=1, value="Campaign snapshot").font = SECTION_FONT
ov.cell(row=snap, column=5, value="(live from Planner →)").font = Font(italic=True, size=9, color="71717A")
snap_hdr = snap + 1
for c, h in enumerate(["Metric", "Value", "", "Metric", "Value"], 1):
    cell = ov.cell(row=snap_hdr, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = NEUTRAL_HEADER
    cell.border = GRID
    cell.alignment = CENTER

base_sum = f"SUM(Planner!D{planner_start}:D{tot_r - 1})"
metrics_left = [
    ("Target budget", "=Planner!B4", FMT_USD),
    ("Global Slot ×", "=Planner!E4", "0.00×"),
    ("Base escrow (×1)", f"={base_sum}", FMT_USD),
]
metrics_right = [
    ("Scaled jobs", f"=Planner!F{tot_r}", FMT_INT),
    ("Scaled escrow", f"=Planner!G{tot_r}", FMT_USD),
    ("Gap vs target", f"=Planner!G{gap_r}", FMT_USD),
]
for j, (label, formula, fmt) in enumerate(metrics_left):
    r = snap_hdr + 1 + j
    ov.cell(row=r, column=1, value=label).font = LABEL_FONT
    ov.cell(row=r, column=1).border = GRID
    v = ov.cell(row=r, column=2, value=formula)
    v.font = Font(bold=True, size=11, color="18181B")
    v.number_format = fmt
    v.border = GRID
for j, (label, formula, fmt) in enumerate(metrics_right):
    r = snap_hdr + 1 + j
    ov.cell(row=r, column=4, value=label).font = LABEL_FONT
    ov.cell(row=r, column=4).border = GRID
    v = ov.cell(row=r, column=5, value=formula)
    v.font = Font(bold=True, size=11, color="18181B")
    v.number_format = fmt
    v.border = GRID
    if label == "Gap vs target":
        v.font = Font(bold=True, size=11, color="B45309")

nav = snap_hdr + 5
ov.cell(row=nav, column=1, value="Where to go").font = SECTION_FONT
nav_hdr = nav + 1
for c, h in enumerate(["Tab", "Bucket", "Use for"], 1):
    cell = ov.cell(row=nav_hdr, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = NEUTRAL_HEADER
    cell.alignment = CENTER
    cell.border = GRID
nav_rows = [
    ("Planner", "—", "Set target $ · scale Slot × · per-bucket breakdown (source of truth)"),
    ("micro", "MICRO", "Photo / pulse / X — $0.10–0.25"),
    ("small", "SMALL", "Connect smoke · friction · stage notes — $0.30–1.00"),
    ("med", "MED", "Job-loop · referral · video — $1–5"),
    ("large", "LARGE", "Ambassadors · shipped tools — $5–25"),
    ("All jobs", "—", "Flat export · filters · pivot"),
    ("Budget", "—", "Static $1k / $5k / $10k reference scenarios"),
]
for j, (tab, bucket, purpose) in enumerate(nav_rows):
    r = nav_hdr + 1 + j
    theme = BUCKET_THEME.get(tab, None)
    row_fill = fill(theme["stripe_a"]) if theme else fill("FAFAFA")
    for c in range(1, 4):
        ov.cell(row=r, column=c).fill = row_fill
        ov.cell(row=r, column=c).border = GRID
    ov.cell(row=r, column=1, value=tab).font = Font(
        bold=True, color=theme["header"] if theme else "18181B"
    )
    ov.cell(row=r, column=2, value=bucket).font = BODY_FONT
    ov.cell(row=r, column=3, value=purpose).font = BODY_FONT
    ov.cell(row=r, column=3).alignment = WRAP

note = nav_hdr + len(nav_rows) + 2
ov.cell(row=note, column=1, value=f"{len(rows_out)} job templates in catalog · edit Max USDC / Batch slots on bucket tabs → Planner updates").font = Font(
    size=10, color="52525B"
)
ov.merge_cells(start_row=note, start_column=1, end_row=note, end_column=5)

for letter, w in [("A", 22), ("B", 14), ("C", 48), ("D", 18), ("E", 14), ("F", 8)]:
    ov.column_dimensions[letter].width = w

# Tab order: Overview · Planner · buckets · Budget · All jobs
wb.move_sheet(pl, 1 - wb.index(pl))

wb.save(OUT_XLSX)

print(f"Wrote {len(rows_out)} templates | {total_slots} slots | ${total_escrow:,.2f} max escrow (base)")
print(f"  CSV:  {OUT_CSV.name}")
print(f"  XLSX: {OUT_XLSX.name} (Planner tab = dynamic budget)")
